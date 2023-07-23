import openpyxl
import pandas as pd
import os


#Carregando planilha Excel
def load_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    else:
        return "arquivo não encontrado"
    
wb_path = "colaboradores.xlsx"
wb = load_workbook(wb_path)
sheet = wb["Sheet"]
sheet_obj = wb.active
max_column = sheet_obj.max_column
max_row = sheet_obj.max_row

#Criar colaborador
def Create_Employee():
    new_employee = input("\n Digite os dados do colaborador: Nome Nascimento E-mail: ").split(' ')
    sheet.append(new_employee)
    wb.save(wb_path)
    print("Detalhes do colaborador adicionados com sucesso!")
    add_more = input("\n Quer adicionar mais colaboradores? s/n ")
    if add_more.lower() == "s":
        Create_Employee()

#Visualizar todos colaboradores
def View_all_employee():
    employee_list = pd.read_excel(wb_path)
    print(employee_list)

#Pesquisar colaborador
def search(nome):
    for i in range(1,max_row+1):
        if sheet.cell(row=i,column = 1).value == nome:
            print("Colaborador encontrado! Seguem os dados:")
            return i
        
#Visualizar colaborador
def Display_employee(row):
    for i in range(1, max_column+1):
        cell_obj = sheet_obj.cell(row = row, column =i)
        print(cell_obj.value)

#Atualizar informações do colaborador
def Update_employee(row):
    x = input("\n Insira os dados do colaborador: Nome Nascimento E-mail :").split(' ')
    for col_index,value in enumerate(x,start =1):
        sheet.cell(row = row, column =col_index,value =value)
    wb.save(wb_path)
    print("\n Dados do Colaborador atualizadas com sucesso")

#Deletar colaborador
def Delete_employee(row):
    sheet.delete_rows(row)
    wb.save(wb_path)
    print("\n Colaborador deletado com sucesso!")

while True:
    print("\n Sistema de Cadastro de Colaboradores")
    print("\n 1.Criar Colaborador")
    print("\n 2.Visualizar colaboradores")
    print("\n 3.Atualizar dados do colaborador")
    print("\n 4.Deletar colaborador")
    ch = input("\n Digite a opção: ")
    if ch == '1':
        Create_Employee()
    if ch == '2':
        View_all_employee()
    if ch == '3':
        x = input("\n Digite o nome do colaborador: ")
        row = search(x)
        Display_employee(row)
        y = input("\n Quer editar esse colaborador? s/n ")
        if y == 's':
            Update_employee(row)
    if ch == '4':
         x = input("\n Digite o nome do colaborador: ")
         row = search(x)
         Display_employee(row)
         y = input("\n Quer deletar esse colaborador? s/n ")
         if y == 's':
            Delete_employee(row)

    else:
        break;

